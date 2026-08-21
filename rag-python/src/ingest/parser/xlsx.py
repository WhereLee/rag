"""xlsx 解析器：每 sheet → 一个 Markdown 表格节点（data_only 取缓存值，合并单元格取左上值）。

- data_only=True：公式单元格取上次保存的缓存值（无缓存为 None）
- 合并单元格：openpyxl 普通模式非左上角返回 None → 自动取左上值
- 全空行/列剔除；空 sheet 跳过
"""
from __future__ import annotations

from pathlib import Path
from typing import List

import openpyxl

from .base import DocumentNode, ParseError, Parser, rows_to_markdown


class XlsxParser(Parser):
    """Excel 解析：sheet → table 节点（meta 带 sheet 名）。"""

    def __init__(self, max_cells: int = 500_000) -> None:
        self.max_cells = max_cells

    def parse(self, path: Path) -> List[DocumentNode]:
        path = Path(path)
        wb = openpyxl.load_workbook(path, data_only=True)
        source = path.name
        total_cells = sum(ws.max_row * ws.max_column for ws in wb.worksheets)
        if total_cells > self.max_cells:
            wb.close()
            raise ParseError(f"{source}: 单元格数超限（{total_cells} > {self.max_cells}）")
        nodes: List[DocumentNode] = []
        for ws in wb.worksheets:
            rows: List[List[str]] = []
            for row in ws.iter_rows():
                vals = ["" if c.value is None else str(c.value).strip() for c in row]
                if any(v for v in vals):
                    rows.append(vals)
            if not rows:
                continue
            md = rows_to_markdown(rows)
            nodes.append(DocumentNode("table", md,
                                      {"source": source, "sheet": ws.title, "index": len(nodes)}))
        wb.close()
        return nodes
